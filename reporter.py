import io
import logging
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from sender import send_report_email

logger = logging.getLogger(__name__)

IST = timezone(timedelta(hours=5, minutes=30))


def _build_report_text(stats: dict, sent_results: list[dict],
                       applications: list[dict]) -> str:
    now = datetime.now(IST).strftime("%B %d, %Y %I:%M %p IST")
    lines = [
        f"Daily Auto-Apply Report — {now}",
        "=" * 60,
        "",
        "SUMMARY",
        "-" * 40,
        f"  Jobs scraped (all platforms):  {stats.get('total_scraped', 0)}",
        f"  After today filter:            {stats.get('today_jobs', 0)}",
        f"  After tech + remote filter:    {stats.get('filtered_jobs', 0)}",
        f"  After dedup:                   {stats.get('unique_jobs', 0)}",
        f"  Matched to templates:          {stats.get('matched_jobs', 0)}",
        f"  Contacts discovered:           {stats.get('contacts_found', 0)}",
        f"  Emails attempted:              {stats.get('emails_attempted', 0)}",
        f"  Emails sent successfully:      {stats.get('emails_sent', 0)}",
        f"  Emails failed:                 {stats.get('emails_failed', 0)}",
        f"  Emails skipped (bounced):      {stats.get('emails_skipped', 0)}",
        "",
    ]

    per_source = stats.get("per_source", {})
    if per_source:
        lines.append("JOBS PER PLATFORM")
        lines.append("-" * 40)
        for source, count in sorted(per_source.items(), key=lambda x: -x[1]):
            lines.append(f"  {source:30s} {count:5d}")
        lines.append("")

    per_account = stats.get("per_account", {})
    if per_account:
        lines.append("EMAILS PER SMTP ACCOUNT")
        lines.append("-" * 40)
        for account, count in per_account.items():
            lines.append(f"  {account:30s} {count:5d}")
        lines.append("")

    if applications:
        lines.append(f"COMPANIES HIRING — EMAILS QUEUED ({len(applications)})")
        lines.append("-" * 60)
        for i, app in enumerate(applications[:80], 1):
            company = app.get("company_name", "?")
            title = app.get("title", "?")
            email = app.get("contact_email", "?")
            tags = ", ".join(app.get("tags", [])[:3])
            loc = app.get("location", "") or "Remote"
            lines.append(f"  {i:3d}. {company} — {title}")
            lines.append(f"       Email: {email} | Location: {loc}")
            if tags:
                lines.append(f"       Tech: {tags}")
        if len(applications) > 80:
            lines.append(f"\n  ... and {len(applications) - 80} more (see Excel)")
        lines.append("")

    sent_ok = [r for r in sent_results if r.get("status") == "sent"]
    if sent_ok:
        lines.append(f"EMAILS SENT ({len(sent_ok)})")
        lines.append("-" * 40)
        for i, r in enumerate(sent_ok[:50], 1):
            company = r.get("company_name", "?")
            email = r.get("email", "?")
            lines.append(f"  {i:3d}. {company} → {email}")
        if len(sent_ok) > 50:
            lines.append(f"  ... and {len(sent_ok) - 50} more (see Excel)")
        lines.append("")

    failed = [r for r in sent_results if r.get("status") == "failed"]
    if failed:
        lines.append(f"FAILED ({len(failed)})")
        lines.append("-" * 40)
        for r in failed[:20]:
            lines.append(f"  {r.get('email', '?')} — {r.get('error', 'unknown')}")
        lines.append("")

    return "\n".join(lines)


def _auto_width(ws):
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)


def _build_excel(applications: list[dict], sent_results: list[dict],
                 all_scraped_jobs: list[dict] | None = None) -> bytes:
    wb = Workbook()

    # Sheet 1: Applications
    ws = wb.active
    ws.title = "Applications"
    headers = ["#", "Company", "Job Title", "Contact Email", "Contact Name",
               "Confidence", "Status", "Location", "Job URL", "Tags"]
    ws.append(headers)

    result_map = {r.get("email", ""): r for r in sent_results}
    for i, app in enumerate(applications, 1):
        result = result_map.get(app.get("contact_email", ""), {})
        status = result.get("status", "pending")
        tags = ", ".join(app.get("tags", [])[:5])
        ws.append([
            i,
            app.get("company_name", ""),
            app.get("title", ""),
            app.get("contact_email", ""),
            app.get("contact_name", "") or app.get("contact_title", ""),
            round(app.get("contact_confidence", 0), 2),
            status,
            app.get("location", "") or "Remote",
            app.get("url", ""),
            tags,
        ])
    _auto_width(ws)

    # Sheet 2: All Jobs (for manual review)
    if all_scraped_jobs:
        ws2 = wb.create_sheet("All Jobs")
        ws2.append(["#", "Company", "Job Title", "Location", "Tags",
                     "Salary", "Apply URL", "Source", "Posted"])
        for i, job in enumerate(all_scraped_jobs[:2000], 1):
            salary = ""
            if job.get("salary_min") and job.get("salary_max"):
                salary = f"${job['salary_min']:,} - ${job['salary_max']:,}"
            elif job.get("salary_min"):
                salary = f"${job['salary_min']:,}+"
            ws2.append([
                i,
                job.get("company_name", ""),
                job.get("title", ""),
                job.get("location", "") or "Remote",
                ", ".join(job.get("tags", [])[:5]),
                salary,
                job.get("url", ""),
                job.get("source", ""),
                str(job.get("posted_at", ""))[:10],
            ])
        _auto_width(ws2)

    # Sheet 3: Failed Sends
    failed = [r for r in sent_results if r.get("status") == "failed"]
    if failed:
        ws3 = wb.create_sheet("Failed Sends")
        ws3.append(["Email", "Company", "Error", "SMTP Account"])
        for r in failed:
            ws3.append([
                r.get("email", ""),
                r.get("company_name", ""),
                r.get("error", ""),
                r.get("via", ""),
            ])
        _auto_width(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def send_daily_report(stats: dict, sent_results: list[dict],
                      applications: list[dict],
                      all_scraped_jobs: list[dict] | None = None):
    today = datetime.now(IST).strftime("%Y-%m-%d")
    sent_n = stats.get("emails_sent", 0)
    fail_n = stats.get("emails_failed", 0)
    scraped_n = stats.get("total_scraped", 0)
    subject = f"Auto-Apply Report — {today} | {sent_n} sent, {fail_n} failed, {scraped_n} scraped"

    body = _build_report_text(stats, sent_results, applications)

    try:
        excel_bytes = _build_excel(applications, sent_results, all_scraped_jobs)
    except Exception as e:
        logger.error("Failed to build Excel report: %s", e)
        excel_bytes = None

    logger.info("Sending daily report...")
    try:
        send_report_email(
            subject=subject,
            body=body,
            attachment_bytes=excel_bytes,
            attachment_name=f"auto_apply_report_{today}.xlsx",
        )
    except Exception as e:
        logger.error("Failed to send report email: %s", e)
        # Try sending without attachment as fallback
        try:
            send_report_email(subject=subject, body=body)
        except Exception as e2:
            logger.error("Fallback report also failed: %s", e2)
